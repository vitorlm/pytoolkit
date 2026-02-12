import os
import re
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

from domains.syngenta.team_assessment.processors.criteria_processor import CriteriaProcessor
from domains.syngenta.team_assessment.services.member_analyzer import MemberAnalyzer
from domains.syngenta.team_assessment.services.team_analyzer import TeamAnalyzer
from utils.data.json_manager import JSONManager
from utils.file_manager import FileManager
from utils.logging.logging_manager import LogManager
from utils.string_utils import StringUtils

from .core.config import Config

# from .services.feedback_specialist import FeedbackSpecialist
from .core.member import Member
from .core.member_productivity_metrics import EpicTaskData, TasksData
from .processors.feedback_processor import FeedbackProcessor
from .processors.members_task_processor import MembersTaskProcessor
from .services.epic_enrichment_service import EpicEnrichmentService
from .services.feedback_analyzer import FeedbackAnalyzer
from .services.historical_period_discovery import HistoricalPeriodDiscovery
from .services.kudos_service import KudosService
from .services.valyou_service import ValYouService
from .services.workday_service import WorkdayService


class AssessmentGenerator:
    """Handles the processing of competency matrix and planning data.

    Generates detailed feedback and assessment reports for team members and teams.
    """

    _logger = LogManager.get_instance().get_logger("AssessmentGenerator")

    def __init__(
        self,
        competency_matrix_file: str,
        feedback_folder: str | None,
        planning_file: str | None,
        output_path: str,
        ignored_member_list: str | None = None,
        enable_historical: bool = True,
        member_slack_mapping: str | None = None,
        enable_kudos: bool = True,
        valyou_file: str | None = None,
    ):
        self.competency_matrix_file = competency_matrix_file
        self.feedback_folder = feedback_folder
        self.planning_file = planning_file
        self.output_path = output_path
        self.enable_historical = enable_historical
        self.enable_kudos = enable_kudos
        self.member_slack_mapping_file = member_slack_mapping
        self.valyou_file = valyou_file

        # Initialize processors
        self.criteria_processor = CriteriaProcessor()
        if planning_file:
            self.task_processor = MembersTaskProcessor()
        if feedback_folder:
            self.feedback_processor = FeedbackProcessor()
        self.config = Config()
        self.feedback_analyzer = FeedbackAnalyzer()
        self.epic_enrichment_service = EpicEnrichmentService()

        # Load ignored members list
        self.ignored_member_list = JSONManager.read_json(ignored_member_list, default=[]) if ignored_member_list else []

        # Initialize member storage
        self.members = {}

        # Initialize historical period discovery
        self.period_discovery = HistoricalPeriodDiscovery()
        self.historical_periods = []
        self.current_period = None

        if self.enable_historical and self.feedback_folder:
            self._discover_periods()

        # Initialize kudos service (if enabled and mapping exists)
        # Must be after period discovery so current_period is available
        self.kudos_service: KudosService | None = None
        if self.enable_kudos and self.config.slack_kudos_enabled:
            self._init_kudos_service()

        # Initialize Val-You service (if CSV provided and enabled)
        self.valyou_service: ValYouService | None = None
        if self.valyou_file and self.config.valyou_enabled:
            self.valyou_service = ValYouService()
            self._logger.info(f"ValYouService initialized for file: {self.valyou_file}")

        # Initialize Workday service
        self.workday_service = WorkdayService()
        self.workday_folder: str | None = None

    def run(self):
        """Executes the process to generate the assessment report."""
        self._logger.info("Starting assessment generation process.")
        self._validate_input_folders()
        self._detect_workday_folder()

        # Process planning data if provided
        if self.planning_file:
            self._process_tasks()
            # Calculate productivity metrics after processing tasks
            self._calculate_productivity_metrics()

        # Process feedback only if feedback_folder is provided
        if self.feedback_folder:
            # Load competency matrix (needed for feedback analysis)
            competency_matrix = self._load_competency_matrix()

            # Process feedback with historical data if enabled
            if self.enable_historical and (self.current_period or self.historical_periods):
                self._logger.info(
                    f"Historical mode enabled. Current period: {self.current_period}, "
                    f"Historical periods: {len(self.historical_periods)}"
                )
                feedback_with_history = self._process_feedback_with_history()
                feedback = feedback_with_history["current"]["data"] if feedback_with_history["current"] else {}
                historical_feedback = feedback_with_history["historical"]

                self._logger.info(f"Processed {len(historical_feedback)} historical periods")
                self._logger.info(f"Current period feedback has {len(feedback)} evaluatees")
                # Store historical data for later use in evolution calculations
                self.historical_feedback_data = historical_feedback
            else:
                self._logger.info("Historical mode disabled - processing current period only")
                feedback = self._process_feedback()
                self.historical_feedback_data = []

            self._logger.info(f"Analyzing feedback for {len(feedback)} evaluatees")
            self._update_members_with_feedback(feedback)

            # Fetch kudos from Slack (after members are populated, before output)
            self._fetch_kudos()
            # Fetch Val-You recognitions (after members are populated, before output)
            self._fetch_valyou_recognitions()
            # Fetch Workday data (after members are populated, before output)
            self._fetch_workday_data()

            team_stats, members_stats = self.feedback_analyzer.analyze(competency_matrix, feedback)
            self._logger.info(
                f"Analysis complete. Team stats criteria: {len(team_stats.criteria_stats)}, Members: {len(members_stats)}"
            )
            self._update_members_with_stats(members_stats)

            # Add historical context to member stats if available
            if self.historical_feedback_data:
                self._add_historical_context_to_stats(members_stats)

            for member_name, member_data in members_stats.items():
                if self._is_member_ignored(member_name):
                    continue
                # Calculate current period label
                current_period_label = (
                    f"{self.current_period.period_name}/{self.current_period.year}"
                    if self.current_period
                    else "Current"
                )
                # Retrieve raw feedback and productivity metrics from Member object
                first_name = member_name.split(" ", 1)[0]
                member_obj = self.members.get(first_name)
                member_raw_feedback = None
                member_productivity = None
                if member_obj:
                    member_raw_feedback = member_obj.feedback
                    member_productivity = (
                        (
                            member_obj.productivity_metrics.model_dump()
                            if hasattr(member_obj.productivity_metrics, "model_dump")
                            else member_obj.productivity_metrics
                        )
                        if member_obj.productivity_metrics
                        else None
                    )

                member_analyzer = MemberAnalyzer(
                    member_name,
                    member_data,
                    team_stats,
                    self.output_path,
                    current_period_label=current_period_label,
                    raw_feedback=member_raw_feedback,
                    productivity_metrics=member_productivity,
                )
                member_analyzer.plot_all_charts()

            # Pass historical data to team analyzer for temporal charts
            historical_stats = getattr(self, "historical_team_stats", [])
            current_period_label = (
                f"{self.current_period.period_name}/{self.current_period.year}" if self.current_period else "Current"
            )
            team_analyzer = TeamAnalyzer(
                team_stats,
                self.output_path,
                historical_stats=historical_stats,
                current_period_label=current_period_label,
            )
            team_analyzer.plot_all_charts()

            self._generate_output(team_stats)
        else:
            self._logger.info("No feedback folder provided - skipping feedback analysis")
            # Fetch kudos even without feedback (members come from planning)
            self._fetch_kudos()
            # Fetch Val-You recognitions even without feedback
            self._fetch_valyou_recognitions()
            # Fetch Workday data even without feedback
            self._fetch_workday_data()
            # Generate output with task data only (if available)
            if self.planning_file:
                self._generate_output_planning_only()
            else:
                self._logger.warning("No feedback folder or planning file provided - nothing to generate")

    def _discover_periods(self):
        """Discovers all evaluation periods (current + historical) from feedback folder."""
        self._logger.info("Discovering evaluation periods...")
        all_periods = self.period_discovery.discover_periods(self.feedback_folder)

        self.current_period = next((p for p in all_periods if p.is_current), None)
        self.historical_periods = [p for p in all_periods if not p.is_current]

        self._logger.info(f"Current period: {self.current_period}")
        self._logger.info(f"Historical periods: {[str(p) for p in self.historical_periods]}")

    def _init_kudos_service(self):
        """Initializes the KudosService if configuration and mapping are available."""
        # Resolve mapping file path
        mapping_file = self.member_slack_mapping_file
        if not mapping_file:
            # Default to member_slack_mapping.json in the team_assessment directory
            default_path = Path(__file__).parent / "member_slack_mapping.json"
            mapping_file = str(default_path)

        if not Path(mapping_file).exists():
            self._logger.warning(f"Member-Slack mapping file not found: {mapping_file} — kudos integration disabled")
            return

        # Determine the assessment year
        year = self._resolve_assessment_year()

        try:
            self.kudos_service = KudosService(
                mapping_file=mapping_file,
                channel=self.config.slack_kudos_channel,
                year=year,
                bot_token=self.config.slack_bot_token,
            )
            self._logger.info(f"KudosService initialized for year {year}")
        except Exception as e:
            self._logger.warning(f"Failed to initialize KudosService: {e} — kudos integration disabled")
            self.kudos_service = None

    def _resolve_assessment_year(self) -> int:
        """Resolves the assessment year from current_period or defaults to current year.

        Returns:
            The assessment year as an integer.
        """
        if self.current_period and hasattr(self.current_period, "year"):
            return int(self.current_period.year)
        return datetime.now(tz=UTC).year

    def _fetch_kudos(self):
        """Fetches kudos for all members from Slack and updates their Member objects.

        Requires that self.members is already populated (called after feedback processing).
        """
        if not self.kudos_service:
            self._logger.debug("KudosService not available — skipping kudos fetch")
            return

        # Build full member names from current members
        member_full_names: list[str] = []
        for member in self.members.values():
            if self._is_member_ignored(member.name):
                continue
            full_name = f"{member.name} {member.last_name}".strip() if member.last_name else member.name
            member_full_names.append(full_name)

        if not member_full_names:
            self._logger.warning("No members available for kudos fetch")
            return

        self._logger.info(f"Fetching kudos for {len(member_full_names)} members...")
        kudos_results = self.kudos_service.fetch_kudos_for_all_members(member_full_names)

        # Update each Member object with kudos data
        for full_name, member_kudos in kudos_results.items():
            # Find member by first name (members dict is keyed by first name)
            first_name = full_name.split(" ", 1)[0]
            if first_name in self.members:
                self.members[first_name].kudos = member_kudos
                if member_kudos.total_count > 0:
                    self._logger.info(
                        f"  {full_name}: {member_kudos.total_count} kudos from {len(member_kudos.senders)} senders"
                    )

        total_kudos = sum(mk.total_count for mk in kudos_results.values())
        self._logger.info(f"Kudos fetch complete: {total_kudos} total across {len(kudos_results)} members")

    def _fetch_valyou_recognitions(self):
        """Fetches Val-You recognitions and updates Member objects.

        Requires that self.members is already populated.
        """
        if not self.valyou_service:
            self._logger.debug("ValYouService not available — skipping Val-You fetch")
            return

        if not self.members:
            self._logger.warning("No members available for Val-You recognition fetch")
            return

        # Filter out ignored members before sending to service
        active_members = {k: v for k, v in self.members.items() if not self._is_member_ignored(v.name)}

        try:
            results = self.valyou_service.fetch_recognitions(self.valyou_file, active_members)
        except Exception as e:
            self._logger.error(f"Failed to fetch Val-You recognitions: {e}", exc_info=True)
            return

        # Update each Member object with Val-You data
        for member_name, recognitions in results.items():
            if member_name in self.members:
                self.members[member_name].valyou_recognitions = recognitions
                if recognitions.total_count > 0:
                    self._logger.info(
                        f"  {member_name}: {recognitions.total_count} Val-You recognitions "
                        f"from {len(recognitions.senders)} senders"
                    )

        total = sum(r.total_count for r in results.values())
        self._logger.info(f"Val-You fetch complete: {total} total across {len(results)} members")

    def _detect_workday_folder(self):
        """Detects the Workday folder relative to the feedback folder.

        The Workday folder is expected to be a sibling of the month folder
        at the YEAR level (e.g., .../2025/Workday).
        """
        if not self.feedback_folder:
            return

        feedback_path = Path(self.feedback_folder)
        # Scan for 'Workday' folder in the parent (YEAR) directory
        year_path = feedback_path.parent
        workday_path = year_path / "Workday"

        if workday_path.exists() and workday_path.is_dir():
            self.workday_folder = str(workday_path)
            self._logger.info(f"Detected Workday folder: {self.workday_folder}")
        else:
            self._logger.debug(f"Workday folder not found as sibling of: {self.feedback_folder}")

    def _fetch_workday_data(self):
        """Fetches Workday data and updates Member objects.

        Requires that self.members is already populated.
        """
        if not self.workday_folder:
            self._logger.debug("Workday folder not detected — skipping Workday fetch")
            return

        if not self.members:
            self._logger.warning("No members available for Workday data fetch")
            return

        # Call service to process the folder
        try:
            results = self.workday_service.process_workday_folder(self.workday_folder, self.members)
        except Exception as e:
            self._logger.error(f"Failed to fetch Workday data: {e}", exc_info=True)
            return

        # Update each Member object with Workday data
        for member_name, workday_data in results.items():
            if member_name in self.members:
                self.members[member_name].workday_data = workday_data
                self._logger.info(f"  {member_name}: Workday data processed")

        self._logger.info(f"Workday fetch complete: {len(results)} members processed")

    def _process_feedback_with_history(self):
        """Processes feedback from current and historical periods.

        Returns a dict with current and historical feedback data.
        """
        result = {"current": None, "historical": []}

        # Process current period
        if self.current_period:
            self._logger.info(f"Processing current period: {self.current_period}")
            period_metadata = {
                "year": self.current_period.year,
                "period_name": self.current_period.period_name,
                "timestamp": self.current_period.timestamp.isoformat(),
            }
            current_data = self.feedback_processor.process_folder(
                self.current_period.folder_path, period_metadata=period_metadata
            )
            result["current"] = {"data": current_data, "period": self.current_period}

        # Process historical periods
        for period in self.historical_periods:
            self._logger.info(f"Processing historical period: {period}")
            period_metadata = {
                "year": period.year,
                "period_name": period.period_name,
                "timestamp": period.timestamp.isoformat(),
            }
            historical_data = self.feedback_processor.process_folder(
                period.folder_path, period_metadata=period_metadata
            )

            result["historical"].append({"data": historical_data, "period": period})

        return result

    def _validate_input_folders(self):
        """Validates the existence and structure of all input folders and files."""
        self._logger.debug("Validating input folders and files.")

        # Validate feedback folder with specific structure requirements (if provided)
        if self.feedback_folder:
            self.period_discovery.validate_feedback_folder_structure(self.feedback_folder)
            self._logger.info(f"Feedback folder validated: {self.feedback_folder}")
        else:
            self._logger.info("No feedback folder provided - skipping feedback validation")

        # Validate competency matrix file
        self._validate_competency_matrix_file()

        # Validate planning file if provided
        if self.planning_file:
            self._validate_planning_file()

        # Validate output folder
        self._validate_output_folder()

    def _validate_output_folder(self):
        """Validates that the output path is a valid folder path (not a file)."""
        # Check if path ends with common file extensions
        file_extensions = [".json", ".xlsx", ".csv", ".txt"]
        if any(self.output_path.endswith(ext) for ext in file_extensions):
            raise ValueError(
                f"Invalid output path: {self.output_path}\n\n"
                f"The --outputFolder parameter must be a folder path, not a file path.\n"
                f"Found file extension in path.\n\n"
                f"Examples of valid paths:\n"
                f"  output/assessment_2025\n"
                f"  /Users/user/reports/team_assessment\n\n"
                f"The system will create JSON files and charts inside this folder."
            )

        # Create output folder if it doesn't exist
        output_dir = Path(self.output_path)
        if not output_dir.exists():
            self._logger.info(f"Creating output folder: {self.output_path}")
            output_dir.mkdir(parents=True, exist_ok=True)
        elif not output_dir.is_dir():
            raise ValueError(
                f"Invalid output path: {self.output_path}\n\n"
                f"Path exists but is not a directory.\n"
                f"Please specify a valid folder path."
            )

        self._logger.info(f"Output folder validated: {self.output_path}")

    def _validate_competency_matrix_file(self):
        """Validates the competency matrix file exists and has correct structure.

        Expected structure:
        - Sheet name: "Competencies and Levels"
        - Column 0: Criterion name
        - Column 1: Indicator name
        - Columns 2-6: Levels 1-5 (descriptions and evidence)
        """
        # Validate file exists and has correct extension
        FileManager.validate_file(self.competency_matrix_file, allowed_extensions=[".xlsx"])
        self._logger.info(f"Competency matrix file found: {self.competency_matrix_file}")

        # Validate sheet structure
        try:
            workbook = openpyxl.load_workbook(self.competency_matrix_file, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames

            # Check for required sheet
            required_sheet = "Competencies and Levels"
            if required_sheet not in sheet_names:
                raise ValueError(
                    f"Invalid competency matrix file structure: {self.competency_matrix_file}\n\n"
                    f"Missing required sheet: '{required_sheet}'\n"
                    f"Found sheets: {', '.join(sheet_names)}\n\n"
                    f"The competency matrix file must contain a sheet named '{required_sheet}' with:\n"
                    f"  - Column A: Criterion name\n"
                    f"  - Column B: Indicator name\n"
                    f"  - Columns C-G: Levels 1-5 (descriptions)"
                )

            # Check minimum structure (at least 2 rows: header + 1 data row)
            sheet = workbook[required_sheet]
            if sheet.max_row < 2:
                raise ValueError(
                    f"Invalid competency matrix file structure: {self.competency_matrix_file}\n\n"
                    f"Sheet '{required_sheet}' appears to be empty.\n"
                    f"Expected at least 2 rows (header + data)"
                )

            # Check minimum columns (Criterion, Indicator, + 5 levels = 7 columns)
            if sheet.max_column < 7:
                raise ValueError(
                    f"Invalid competency matrix file structure: {self.competency_matrix_file}\n\n"
                    f"Sheet '{required_sheet}' has insufficient columns.\n"
                    f"Found: {sheet.max_column} columns\n"
                    f"Expected: At least 7 columns (Criterion, Indicator, Levels 1-5)"
                )

            workbook.close()
            self._logger.info("Competency matrix structure validated successfully")

        except ImportError:
            self._logger.warning("openpyxl not available - skipping detailed structure validation")
        except Exception as e:
            if isinstance(e, ValueError):
                raise
            raise ValueError(
                f"Error validating competency matrix file structure: {self.competency_matrix_file}\n\nError: {e!s}"
            ) from e

    def _validate_planning_file(self):
        """Validates the planning file exists and has correct structure.

        Expected structure:
        - At least one sheet matching pattern Q[1-4]-C[1-2]
        - Configured columns for members, tasks, allocation matrix
        - Data in expected row ranges
        """
        if not self.planning_file:
            raise ValueError("Planning file path is required but was not provided")

        # Validate file exists and has correct extension
        FileManager.validate_file(self.planning_file, allowed_extensions=[".xlsm", ".xlsx"])
        self._logger.info(f"Planning file found: {self.planning_file}")

        # Validate sheet structure
        try:
            workbook = openpyxl.load_workbook(self.planning_file, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames

            # Check for sheets matching pattern Q[1-4]-C[1-2]
            pattern = r"Q[1-4]-C[1-2]"
            matching_sheets = [name for name in sheet_names if re.match(pattern, name)]

            if not matching_sheets:
                raise ValueError(
                    f"Invalid planning file structure: {self.planning_file}\n\n"
                    f"No sheets found matching pattern '{pattern}'\n"
                    f"Found sheets: {', '.join(sheet_names)}\n\n"
                    f"The planning file must contain at least one sheet with name like:\n"
                    f"  Q1-C1, Q1-C2, Q2-C1, Q2-C2, Q3-C1, Q3-C2, Q4-C1, Q4-C2\n\n"
                    f"Each sheet should contain:\n"
                    f"  - Member list (rows 4-14)\n"
                    f"  - Task backlog (rows 6-33)\n"
                    f"  - Allocation matrix (rows 39-47)"
                )

            # Validate structure of first matching sheet
            sheet = workbook[matching_sheets[0]]

            # Check minimum rows
            if sheet.max_row < 47:
                raise ValueError(
                    f"Invalid planning file structure: {self.planning_file}\n\n"
                    f"Sheet '{matching_sheets[0]}' has insufficient rows.\n"
                    f"Found: {sheet.max_row} rows\n"
                    f"Expected: At least 47 rows (for allocation matrix ending at row 47)"
                )

            # Check minimum columns (allocation matrix starts at column K)
            min_columns = ord("K") - ord("A") + 1  # At least up to column K
            if sheet.max_column < min_columns:
                raise ValueError(
                    f"Invalid planning file structure: {self.planning_file}\n\n"
                    f"Sheet '{matching_sheets[0]}' has insufficient columns.\n"
                    f"Found: {sheet.max_column} columns\n"
                    f"Expected: At least {min_columns} columns (up to column K for allocation)"
                )

            workbook.close()
            self._logger.info(
                f"Planning file structure validated successfully. "
                f"Found {len(matching_sheets)} valid sheets: {', '.join(matching_sheets)}"
            )

        except ImportError:
            self._logger.warning("openpyxl not available - skipping detailed structure validation")
        except Exception as e:
            if isinstance(e, ValueError):
                raise
            raise ValueError(f"Error validating planning file structure: {self.planning_file}\n\nError: {e!s}") from e

    def _load_competency_matrix(self):
        """Loads the default competency matrix from the specified file."""
        self._logger.info(f"Loading competency matrix from: {self.competency_matrix_file}")
        return self.criteria_processor.process_file(Path(self.competency_matrix_file))

    def _process_tasks(self):
        """Processes the planning data to extract tasks for each member."""
        if not self.planning_file:
            self._logger.warning("No planning file provided - skipping task processing")
            return

        self._logger.info(f"Processing tasks from file: {self.planning_file}")
        members_tasks = self.task_processor.process_file(self.planning_file)

        if not members_tasks:
            self._logger.warning(f"No task data extracted from planning file: {self.planning_file}")
        else:
            self._logger.info(f"Extracted tasks for {len(members_tasks)} members")

        for member_name, member_tasks in members_tasks.items():
            self._update_member_tasks(member_name, member_tasks)

    def _calculate_productivity_metrics(self):
        """Calculates detailed productivity metrics for all members based on Planning data.

        Analyzes task allocations across 9 categories (EPIC, SUPPORT, BUG, IMPROVEMENT,
        OVERDUE, ABSENCE, LEARNING, PLANNING, UNPLANNED) to provide comprehensive
        productivity insights including:
        - Time distribution by category
        - Focus on value (% time in EPICs)
        - Maintenance workload (Bugs + Support)
        - Absence impact
        - Task details and adherence metrics
        """
        if not self.planning_file:
            self._logger.info("No planning file provided - skipping productivity metrics")
            return

        self._logger.info("Calculating productivity metrics for all members...")

        for member_name, member in self.members.items():
            if self._is_member_ignored(member_name):
                continue

            if not member.tasks:
                self._logger.debug(f"No tasks found for {member_name} - skipping productivity metrics")
                continue

            try:
                metrics = self._calculate_member_productivity(member_name, member.tasks)

                # Separate tasks_data from metrics (new structure)
                tasks_data = metrics.pop("tasks_data", None)
                member.tasks_data = tasks_data

                # Remove tasks_by_category and tasks_by_work_type from metrics (moved to tasks_data)
                metrics.pop("tasks_by_category", None)
                metrics.pop("tasks_by_work_type", None)

                member.productivity_metrics = metrics
                self._logger.info(
                    f"Calculated productivity metrics for {member_name}: "
                    f"{metrics['summary']['total_allocated_days']:.1f} days, "
                    f"{metrics['summary']['value_focus']:.1f}% focus on value"
                )
            except Exception as e:
                self._logger.error(f"Failed to calculate productivity metrics for {member_name}: {e}", exc_info=True)

        calculated_count = len([m for m in self.members.values() if m.productivity_metrics])
        self._logger.info(f"Productivity metrics calculated for {calculated_count} members")

    def _calculate_member_productivity(self, member_name: str, tasks: list) -> dict:
        """Calculates productivity metrics for a single member.

        Args:
            member_name: Name of the member
            tasks: List of Issue objects assigned to the member

        Returns:
            dict: Comprehensive productivity metrics including:
                - summary: High-level metrics (total days, focus, etc.)
                - distribution: Time distribution by category (9 types)
                - tasks_by_category: Detailed task lists per category
                - key_metrics: Calculated KPIs
        """
        # Initialize category tracking (9 categories including IDLE)
        # Note: BUG category removed - all bug tasks go to SUPPORT
        categories = {
            "EPIC": {"days": 0.0, "tasks": []},
            "SUPPORT": {"days": 0.0, "tasks": []},
            "IMPROVEMENT": {"days": 0.0, "tasks": []},
            "OVERDUE": {"days": 0.0, "tasks": []},
            "ABSENCE": {"days": 0.0, "tasks": []},
            "LEARNING": {"days": 0.0, "tasks": []},
            "PLANNING": {"days": 0.0, "tasks": []},
            "UNPLANNED": {"days": 0.0, "tasks": []},
            "IDLE": {"days": 0.0, "tasks": []},
        }

        # Initialize work type tracking (5 types: Eng, Prod, Day-by-Day, Out, Bug)
        work_types = {
            "Eng": {"days": 0.0, "tasks": []},
            "Prod": {"days": 0.0, "tasks": []},
            "Day-by-Day": {"days": 0.0, "tasks": []},
            "Out": {"days": 0.0, "tasks": []},
            "Bug": {"days": 0.0, "tasks": []},
        }

        total_allocated_days = 0.0  # All days including IDLE
        total_working_days = 0.0  # Days in team (excluding IDLE)

        # Process each task
        for task in tasks:
            if not task:
                continue

            # Get category from task (default to EPIC for development work)
            category = getattr(task, "category", "EPIC") or "EPIC"

            # Map DEVELOPMENT category to EPIC for backward compatibility
            if category == "DEVELOPMENT":
                category = "EPIC"

            # Get days allocated to this task
            task_days = 0.0
            if hasattr(task, "executed") and task.executed:
                task_days = float(getattr(task.executed, "issue_total_days", 0))

            total_allocated_days += task_days

            # Count toward working days only if NOT IDLE
            if category != "IDLE":
                total_working_days += task_days

            # Store task details
            if category in categories:
                categories[category]["days"] += task_days

                # For non-EPIC categories, consolidate tasks with same code
                # EPICs can repeat across cycles, but non-epics should be consolidated
                task_code = getattr(task, "code", "")

                if category == "EPIC":
                    # EPICs: Keep all occurrences (can span multiple cycles)
                    categories[category]["tasks"].append(
                        {
                            "code": task_code,
                            "jira_key": getattr(task, "jira", ""),
                            "description": getattr(task, "description", ""),
                            "type": getattr(task, "type", ""),
                            "category": category,
                            "allocated_days": task_days,
                            "status": getattr(task, "status", "Unknown"),
                        }
                    )
                else:
                    # Non-EPICs: Consolidate by code (sum days)
                    existing_task = None
                    for existing in categories[category]["tasks"]:
                        if existing["code"] == task_code:
                            existing_task = existing
                            break

                    if existing_task:
                        # Update existing task with accumulated days
                        existing_task["allocated_days"] += task_days
                    else:
                        # Add new task
                        categories[category]["tasks"].append(
                            {
                                "code": task_code,
                                "jira_key": getattr(task, "jira", ""),
                                "description": getattr(task, "description", ""),
                                "type": getattr(task, "type", ""),
                                "category": category,
                                "allocated_days": task_days,
                                "status": getattr(task, "status", "Unknown"),
                            }
                        )

            # Track work type distribution (only for working days, not IDLE)
            if category != "IDLE":
                work_type = getattr(task, "type", "Day-by-Day") or "Day-by-Day"
                # Normalize work type to match expected values
                if work_type not in work_types:
                    work_type = "Day-by-Day"  # Default for unknown types
                work_types[work_type]["days"] += task_days

                # Apply same consolidation logic as categories
                if category == "EPIC":
                    # EPICs: Keep all occurrences
                    work_types[work_type]["tasks"].append(
                        {
                            "code": task_code,
                            "category": category,
                            "allocated_days": task_days,
                        }
                    )
                else:
                    # Non-EPICs: Consolidate by code
                    existing_wt_task = None
                    for existing in work_types[work_type]["tasks"]:
                        if existing["code"] == task_code and existing["category"] == category:
                            existing_wt_task = existing
                            break

                    if existing_wt_task:
                        existing_wt_task["allocated_days"] += task_days
                    else:
                        work_types[work_type]["tasks"].append(
                            {
                                "code": task_code,
                                "category": category,
                                "allocated_days": task_days,
                            }
                        )

        # Calculate time metrics
        absence_days = categories["ABSENCE"]["days"]
        idle_days = categories["IDLE"]["days"]
        available_days = total_working_days - absence_days  # Working days minus absences

        # Calculate percentages for each category (based on total_working_days to show distribution of actual work)
        distribution = {}
        for cat_name, cat_data in categories.items():
            # Calculate percentage relative to working days (excludes IDLE)
            if cat_name == "IDLE":
                # IDLE percentage relative to total allocated days
                percentage = (cat_data["days"] / total_allocated_days * 100) if total_allocated_days > 0 else 0.0
            else:
                # Other categories relative to working days
                percentage = (cat_data["days"] / total_working_days * 100) if total_working_days > 0 else 0.0

            distribution[cat_name] = {
                "days": round(cat_data["days"], 2),
                "percentage": round(percentage, 2),
                "task_count": len(cat_data["tasks"]),
            }

        # Calculate key metrics (all based on available_days to exclude absences from KPIs)
        epic_days = categories["EPIC"]["days"]
        support_days = categories["SUPPORT"]["days"]
        overdue_days = categories["OVERDUE"]["days"]

        # 1. Value Focus: % time on EPICs (relative to working days minus absences)
        value_focus = (epic_days / available_days * 100) if available_days > 0 else 0.0

        # 2. Maintenance Time: Support only (bugs included in SUPPORT category)
        maintenance_days = support_days
        maintenance_rate = (maintenance_days / available_days * 100) if available_days > 0 else 0.0

        # 3. Spillover/Overdue Rate (relative to working days minus absences)
        spillover_rate = (overdue_days / available_days * 100) if available_days > 0 else 0.0

        # 4. Utilization rate (% of allocated days that are actual working days)
        utilization_rate = (total_working_days / total_allocated_days * 100) if total_allocated_days > 0 else 0.0

        # Calculate work type distribution (percentage relative to working days)
        work_type_distribution = {}
        for type_name, type_data in work_types.items():
            percentage = (type_data["days"] / total_working_days * 100) if total_working_days > 0 else 0.0
            work_type_distribution[type_name] = {
                "days": round(type_data["days"], 2),
                "percentage": round(percentage, 2),
                "task_count": len(type_data["tasks"]),
            }

        # Build comprehensive metrics structure
        metrics = {
            "member_name": member_name,
            "summary": {
                "total_allocated_days": round(total_allocated_days, 2),
                "total_working_days": round(total_working_days, 2),
                "idle_days": round(idle_days, 2),
                "absence_days": round(absence_days, 2),
                "available_days": round(available_days, 2),
                "total_tasks": len(tasks),
                "value_focus": round(value_focus, 2),
                "maintenance_rate": round(maintenance_rate, 2),
                "spillover_rate": round(spillover_rate, 2),
                "utilization_rate": round(utilization_rate, 2),
            },
            "distribution": distribution,
            "work_type_distribution": work_type_distribution,
            "tasks_by_category": {cat: data["tasks"] for cat, data in categories.items() if data["tasks"]},
            "tasks_by_work_type": {wt: data["tasks"] for wt, data in work_types.items() if data["tasks"]},
            "key_metrics": {
                "value_focus": {
                    "value": round(value_focus, 2),
                    "description": "Percentage of time spent on EPICs (value delivery)",
                    "benchmark": ">60% = Excellent, >45% = Good, >30% = Fair",
                },
                "maintenance_rate": {
                    "value": round(maintenance_rate, 2),
                    "description": "Percentage of time on bugs and support (maintenance)",
                    "benchmark": "<25% = Excellent, <35% = Good, <45% = Fair",
                },
                "spillover_rate": {
                    "value": round(spillover_rate, 2),
                    "description": "Percentage of time on overdue work (rework indicator)",
                    "benchmark": "<10% = Excellent, <20% = Good, <30% = Fair",
                },
                "utilization_rate": {
                    "value": round(utilization_rate, 2),
                    "description": "Percentage of available time utilized",
                    "benchmark": ">90% = Excellent, >80% = Good, >70% = Fair",
                },
            },
        }

        # Extract EPIC keys from member's tasks for enrichment
        # Use set to eliminate duplicates (same epic can appear multiple times in planning)
        epic_keys_set = set()
        for task in tasks:
            if not task:
                continue
            category = getattr(task, "category", "EPIC") or "EPIC"
            # Map DEVELOPMENT category to EPIC for backward compatibility
            if category == "DEVELOPMENT":
                category = "EPIC"

            # Get JIRA key from task (try jira attribute first, then code)
            jira_key = getattr(task, "jira", None) or getattr(task, "code", None)
            if category == "EPIC" and jira_key and "-" in jira_key:
                epic_keys_set.add(jira_key)

        # Convert set to list for processing
        epic_keys = list(epic_keys_set)

        # Enrich epics with JIRA data (if any epics found)
        epic_adherence = None
        epic_metrics = []
        epic_adherence_obj = None
        if epic_keys:
            try:
                self._logger.info(f"Enriching {len(epic_keys)} unique epics for {member_name}")

                # Fetch JIRA data for all epics
                enrichment_data = self.epic_enrichment_service.enrich_epics(epic_keys)

                # Calculate adherence metrics for each epic (deduplicated)
                for epic_key in epic_keys:
                    if epic_key in enrichment_data:
                        epic_data = enrichment_data[epic_key]
                        adherence = self.epic_enrichment_service.calculate_adherence(epic_data)
                        epic_metrics.append(adherence)
                    else:
                        self._logger.warning(f"Epic {epic_key} not found in JIRA response")

                # Aggregate epic adherence statistics
                epic_adherence_obj = self.epic_enrichment_service.aggregate_epic_adherence(epic_metrics)
                epic_adherence = epic_adherence_obj.model_dump()

                self._logger.info(
                    f"Epic adherence calculated for {member_name}: "
                    f"{epic_adherence_obj.total_epics} unique epics, "
                    f"{epic_adherence_obj.adherence_rate:.1f}% adherence rate"
                )

            except Exception as e:
                self._logger.error(f"Failed to enrich epics for {member_name}: {e}", exc_info=True)

        # Build TasksData structure (separates raw data from metrics)
        epic_task_data = []
        enrichment_data_map = {}
        if epic_keys:
            try:
                # Get full enrichment data for all epics (includes all 4 dates)
                enrichment_data_map = self.epic_enrichment_service.enrich_epics(epic_keys)
            except Exception as e:
                self._logger.warning(f"Failed to get enrichment data for task data: {e}")

        for task in tasks:
            if not task:
                continue
            category = getattr(task, "category", "EPIC") or "EPIC"
            if category == "DEVELOPMENT":
                category = "EPIC"

            if category == "EPIC":
                jira_key = getattr(task, "jira", None) or getattr(task, "code", None)
                if jira_key and "-" in jira_key:
                    # Find corresponding adherence data and enrichment data
                    adherence_data = None
                    for metric in epic_metrics:
                        if metric.epic_key == jira_key:
                            adherence_data = metric
                            break

                    # Get full enrichment data (includes all 4 dates)
                    enrichment = enrichment_data_map.get(jira_key)

                    # Build task dict
                    task_dict = {
                        "jira_key": jira_key,
                        "description": getattr(task, "description", ""),
                        "allocated_days": float(getattr(task.executed, "issue_total_days", 0))
                        if hasattr(task, "executed") and task.executed
                        else 0.0,
                        "type": getattr(task, "type", ""),
                        "category": category,
                    }

                    epic_data = EpicTaskData.from_task_and_adherence(task_dict, adherence_data, enrichment)
                    epic_task_data.append(epic_data)

        # Sort epics by adherence_status then days_difference
        def epic_sort_key(epic: EpicTaskData):
            """Sort key: on_time/early first, then late, then in_progress, then by days_difference."""
            if not epic.jira_enrichment:
                return (4, 0)  # no_dates last

            status = epic.jira_enrichment.get("adherence_status", "")
            days_diff = epic.jira_enrichment.get("days_difference", 0) or 0

            status_priority = {
                "early": (0, days_diff),  # Most negative first
                "on_time": (1, days_diff),  # 0 or ±1
                "late": (2, days_diff),  # Most positive first
                "in_progress": (3, 0),
                "no_dates": (4, 0),
            }

            return status_priority.get(status, (5, 0))

        epic_task_data.sort(key=epic_sort_key)

        # Build TasksData
        tasks_data = TasksData(
            epics=epic_task_data,
            tasks_by_category={cat: data["tasks"] for cat, data in categories.items() if data["tasks"]},
            tasks_by_work_type={wt: data["tasks"] for wt, data in work_types.items() if data["tasks"]},
        )

        # Calculate completed epics count
        completed_count = 0
        if epic_metrics:
            completed_count = sum(1 for m in epic_metrics if m.adherence_status in ["on_time", "early", "late"])

        # Update summary with epic metrics
        metrics["summary"].update(
            {
                "epic_adherence_rate": epic_adherence_obj.adherence_rate if epic_adherence_obj else 0.0,
                "total_epics": epic_adherence_obj.total_epics if epic_adherence_obj else 0,
                "epics_completed": completed_count,
                "epics_on_time": epic_adherence_obj.on_time if epic_adherence_obj else 0,
                "epics_early": epic_adherence_obj.early if epic_adherence_obj else 0,
                "epics_late": epic_adherence_obj.late if epic_adherence_obj else 0,
                "epics_in_progress": epic_adherence_obj.in_progress if epic_adherence_obj else 0,
            }
        )

        # Rename epic_adherence to epic_adherence_summary (clearer name)
        metrics["epic_adherence_summary"] = epic_adherence

        # Add tasks_data at root level
        metrics["tasks_data"] = tasks_data.model_dump()

        return metrics

    def _process_feedback(self):
        """Processes feedback data and returns the competency matrix."""
        self._logger.info(f"Processing feedback from: {self.feedback_folder}")
        return self.feedback_processor.process_folder(self.feedback_folder)

    def _update_member_tasks(self, member_name, tasks):
        """Updates or initializes member tasks."""
        member_name = StringUtils.remove_accents(member_name)
        if self._is_member_ignored(member_name):
            return

        if member_name not in self.members:
            self.members[member_name] = Member(name=member_name, tasks=list(tasks), feedback=None)
        else:
            self.members[member_name].tasks = list(tasks)

    def _update_members_with_feedback(self, competency_matrix):
        """Update members with feedback from the competency matrix.

        This method processes a competency matrix containing evaluations and updates
        the feedback for each member. It ensures that each member exists in the
        members dictionary and safely handles feedback updates.

        Args:
            competency_matrix (dict): A dictionary where the keys are evaluatee names
                                      (str) and the values are dictionaries containing
                                      evaluator names (str) and their feedback (any).

        Raises:
            ValueError: If the evaluatee name cannot be split into a first name and last name.
        """
        for evaluatee_name, evaluations in competency_matrix.items():
            if self._is_member_ignored(evaluatee_name):
                continue
            # Handle names with or without last name
            name_parts = evaluatee_name.split(" ", 1)
            name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""
            name = StringUtils.remove_accents(name)
            last_name = StringUtils.remove_accents(last_name)

            # Ensure the member exists in the dictionary
            if name not in self.members:
                self.members[name] = Member(
                    name=name,
                    last_name=last_name,
                    feedback={},
                    feedback_stats=None,
                    tasks=[],
                    health_check=None,
                )

            # Update last_name if member exists but has no last_name (e.g., created from planning)
            if self.members[name].last_name is None and last_name:
                self.members[name].last_name = last_name

            # Safely handle feedback updates
            if self.members[name].feedback is None:
                self.members[name].feedback = {}

            for evaluator_name, feedback in evaluations.items():
                evaluator_normalized = StringUtils.remove_accents(evaluator_name)
                self.members[name].feedback[evaluator_normalized] = feedback

    def _add_historical_context_to_stats(self, members_stats):
        """Adds historical evaluation data to member stats for comparison.

        Also analyzes historical data to generate statistics for temporal comparison.
        """
        if not self.historical_feedback_data:
            return

        self._logger.info("Adding historical context to member stats...")

        # Store analyzed historical stats for team-level temporal analysis
        self.historical_team_stats = []

        for historical_entry in self.historical_feedback_data:
            period = historical_entry["period"]
            historical_data = historical_entry["data"]

            # Analyze historical data to generate statistics
            try:
                # Process competency matrix for historical analysis
                competency_matrix = self.criteria_processor.process_file(Path(self.competency_matrix_file))
                hist_team_stats, hist_members_stats = self.feedback_analyzer.analyze(competency_matrix, historical_data)

                # Store historical team stats with period info
                self.historical_team_stats.append(
                    {
                        "period": {
                            "year": period.year,
                            "period_name": period.period_name,
                            "timestamp": period.timestamp.isoformat(),
                            "label": f"{period.period_name}/{period.year}",
                        },
                        "team_stats": hist_team_stats,
                        "members_stats": hist_members_stats,
                    }
                )

            except Exception as e:
                self._logger.warning(f"Failed to analyze historical period {period}: {e}")

            # Add raw historical data to individual member stats
            for evaluatee_name, evaluations in historical_data.items():
                evaluatee_normalized = StringUtils.remove_accents(evaluatee_name)

                if evaluatee_normalized not in members_stats:
                    continue

                # Append to historical_evaluations (now a proper Pydantic field)
                members_stats[evaluatee_normalized].historical_evaluations.append(
                    {
                        "period": {
                            "year": period.year,
                            "period_name": period.period_name,
                            "timestamp": period.timestamp.isoformat(),
                        },
                        "data": evaluations,
                    }
                )

        self._logger.info(f"Added historical context for {len(members_stats)} members")
        self._logger.info(f"Analyzed {len(self.historical_team_stats)} historical periods for temporal comparison")

    def _update_members_with_stats(self, members_stats):
        """Updates members with analyzed feedback stats."""
        for member_name, member_stats in members_stats.items():
            if self._is_member_ignored(member_name):
                continue
            # Handle names with or without last name
            name_parts = member_name.split(" ", 1)
            name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""
            name = StringUtils.remove_accents(name)
            last_name = StringUtils.remove_accents(last_name)

            if name not in self.members:
                self.members[name] = Member(
                    name=name,
                    last_name=last_name,
                    feedback={},
                    feedback_stats=None,
                    tasks=[],
                    health_check=None,
                )

            # Update feedback_stats
            self.members[name].feedback_stats = member_stats

    def _generate_output(self, team_stats):
        """Generates the final output report."""
        self._logger.info(f"Generating output report at {self.output_path}")

        # Create a directory for member outputs
        members_output_path = os.path.join(self.output_path, "members")
        FileManager.create_folder(members_output_path)

        # Store each member's data in a separate file
        for member in self.members.values():
            member_data = {
                "name": member.name,
                "health_check": member.health_check,
                "feedback": member.feedback,
                "feedback_stats": member.feedback_stats,
                "tasks_data": member.tasks_data,  # NEW: Separated task data
                "productivity_metrics": member.productivity_metrics,
                "kudos": member.kudos.model_dump() if member.kudos else None,
                "valyou_recognitions": member.valyou_recognitions.model_dump() if member.valyou_recognitions else None,
                "workday": member.workday_data.model_dump() if member.workday_data else None,
                # REMOVED: "tasks" array (fully duplicated in tasks_data)
            }
            member_output_folder = os.path.join(members_output_path, member.name.split()[0])
            FileManager.create_folder(member_output_folder)
            member_file_path = os.path.join(member_output_folder, "stats.json")
            JSONManager.write_json(member_data, member_file_path)

        # Store team stats in a separate file
        team_file_path = os.path.join(self.output_path, "team_stats.json")
        JSONManager.write_json({"team": team_stats}, team_file_path)

        self._logger.info("Assessment report successfully generated.")

    def _generate_output_planning_only(self):
        """Generates output report with planning data only (no feedback)."""
        self._logger.info(f"Generating planning-only output report at {self.output_path}")

        # Create a directory for member outputs (same structure as full assessment)
        members_output_path = os.path.join(self.output_path, "members")
        FileManager.create_folder(members_output_path)

        # Store each member's data in a separate file
        for member in self.members.values():
            if self._is_member_ignored(member.name):
                continue

            member_data = {
                "name": member.name,
                "last_name": member.last_name,
                "tasks_data": member.tasks_data,  # Task data from planning
                "productivity_metrics": member.productivity_metrics,  # Productivity metrics
                "kudos": member.kudos.model_dump() if member.kudos else None,
                "valyou_recognitions": member.valyou_recognitions.model_dump() if member.valyou_recognitions else None,
                "workday": member.workday_data.model_dump() if member.workday_data else None,
                # No feedback fields since we don't have feedback data
            }
            member_output_folder = os.path.join(members_output_path, member.name.split()[0])
            FileManager.create_folder(member_output_folder)
            member_file_path = os.path.join(member_output_folder, "planning_stats.json")
            JSONManager.write_json(member_data, member_file_path)
            self._logger.info(f"Planning output generated for {member.name}: {member_file_path}")

        self._logger.info("Planning-only report successfully generated.")

    def _is_member_ignored(self, member_name):
        """Checks if a member is in the ignored member list.

        Args:
            member_name (str): The name of the member to check.

        Returns:
            bool: True if the member is in the ignored member list, False otherwise.
        """
        name_parts = member_name.split(" ", 1)
        first_name = StringUtils.remove_accents(name_parts[0])
        last_name = StringUtils.remove_accents(name_parts[1]) if len(name_parts) > 1 else ""

        for ignored_member in self.ignored_member_list:
            ignored_parts = ignored_member.split(" ", 1)
            ignored_first_name = StringUtils.remove_accents(ignored_parts[0])
            ignored_last_name = StringUtils.remove_accents(ignored_parts[1]) if len(ignored_parts) > 1 else ""

            if first_name == ignored_first_name:
                if not last_name or last_name == ignored_last_name:
                    return True

        return False
